import streamlit as st
from io import StringIO
import calendar
import csv
import io
import re
import pdfplumber
from openpyxl import load_workbook
import xlrd ,openpyxl
import pandas as pd
from io import BytesIO

class DatamapperDocumentParser:
    def __init__(self):
            self.line_list = []
            self.header_index = ['day', 'Day', 'date', 'Date', 'DATE', 'DAY']
            self.name_index = ['name','Name','NAME','Company_name','COMPANY_NAME','CompanyName']
            self.month_index = [
            "January", "February", "March", "April", "May", "June", 
            "July", "August", "September", "October", "November", "December",
            "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
                ]
            self.lower_month_index = [x.lower() for x in self.month_index]
            self.upper_month_index = [x.upper() for x in self.month_index]
            self.full_month_index = self.month_index + self.lower_month_index + self.upper_month_index
            self.month = {month: index for index, month in enumerate(calendar.month_name) if month}
            self.header_index = ['day', 'Day', 'date', 'Date', 'DATE', 'DAY']

    
    
    def extract_name_meter(self,text):
        meter_value = None
        name_value = None
        for part in text:
            

            if 'Meter #' in part:
                    meter_field = part.split('Meter #:')[1].strip()
                    if meter_field:
                        meter_value = meter_field  # Assign the meter value
                    else:
                        meter_value = None  # In case the meter field is empty
                        
            if 'Name:' in part:
                    name_field = part.split('Name:')[1].strip()
                    if name_field == "" or name_field == "None":
                        name_value = None  # Treat the name as None if empty or invalid
                    else:
                        name_value = name_field  # Extract the valid name

        return meter_value, name_value


    def extract_name_and_meter_excel(self,data):
        meter_value = None
        name_value = None
        
        # Split the data into words or sections to process them
        sections = data.split('None')
        
        # Loop through the sections to extract the Meter and Name information
        for part in sections:
            part = part.strip()  # Clean up extra spaces and newlines
            
            # Look for "Meter #" and "Name:"
            if 'Meter #' in part:
                meter_field = part.split('Meter #:')[1].strip()
                if meter_field:
                    meter_value = meter_field  # Assign the meter value
                else:
                    meter_value = None  # In case the meter field is empty
                    
            if 'Name:' in part:
                name_field = part.split('Name:')[1].strip()
                if name_field == "" or name_field == "None":
                    name_value = None  # Treat the name as None if empty or invalid
                else:
                    name_value = name_field  # Extract the valid name

        return meter_value, name_value


    def extract_month(self,text):
        month = None
        
        month_match = re.findall(r'([A-Za-z]+)\s+(\d{4})', text)
        
        if month_match:
            month = month_match
        else:
            print("Month not found in the text")
        
        return month      


    def convert_single_to_multi_sheet_xlsx(self, input_xls_file, output_file_name='converted_file.xlsx'):
    # Load the .xls file
        xls = pd.ExcelFile(input_xls_file)
        
        # Check if the file is already multi-sheet
        if len(xls.sheet_names) > 1:
            print("The file is already a multi-sheet Excel file. No conversion needed.")
            return {"status": True, "file": input_xls_file}

        # Read the single sheet into a DataFrame
        df = pd.read_excel(input_xls_file)

        # Create a new Excel writer object with openpyxl engine
        with pd.ExcelWriter(output_file_name, engine='openpyxl') as writer:
            # Save the single sheet as 'Sheet1' in the new .xlsx file
            df.to_excel(writer, sheet_name='Sheet1', index=False)

            # Add additional empty sheets to make it multi-sheet
            pd.DataFrame().to_excel(writer, sheet_name='Sheet2', index=False)
            pd.DataFrame().to_excel(writer, sheet_name='Sheet3', index=False)

        print(f"Conversion completed: The file has been converted to a multi-sheet Excel file: {output_file_name}")
        return {"status": True, "file": output_file_name}



    def parse_csv_from_pdf(self , pdf_file):
        csv_file = 'output.csv'
        #line_list = []
        csv_output = StringIO()
        buff = StringIO()
        with pdfplumber.open(pdf_file) as pdf:
            csvwriter = csv.writer(csv_output)
            header_written = False

            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    print("No text found on this page")
                    continue

                lines = text.splitlines()
                meter,name = self.extract_name_meter(lines)
                
                print("LINES>>",lines)
                if len(lines)  > 1:
                    day_month = lines[1]

                table_started = False
                for line in lines:
                    if any(word in line for word in self.header_index):
                        self.line_list.append(line)

                for line in lines:
                    if any(word in line for word in self.header_index):
                        
                        if not header_written:
                            header_row = ['Name', 'MeterID'] + line.split()
                            csvwriter.writerow(header_row)
                            header_written = True
                        table_started = True

                    if table_started:
                        row = line.split()
                        try:
                            p = int(row[0])
                            if row[3:]:
                                date = f"{self.month.get(day_month.split()[0])}/{row[0]}/{day_month.split()[1]}"
                                row[0] = date
                                csvwriter.writerow([name if name else '', meter if meter else ''] + row)
                        except:
                            continue

        csv_output.seek(0)
        print("s",csv_output)
        
        df = pd.read_csv(csv_output)
        df.columns = [
    "Name",
    "MeterID",
    "Day",
    "Differential (In. H2O)",
    "Pressure (psig)",
    "Temp. (°F)",
    "Flow Time (hrs)",
    "Relative Density",
    "Plate (inches)",
    "Volume (Mcf)",
    "Heating Value (Btu/scf)",
    "Energy (MMBtu)"
    
]
        print("DF>>" , df)
        csv_output = df.to_csv(buff,index=False)
        buff.seek(0)
        print("bbb",csv_output)
      
        return buff, csv_file ,df
    

    def parse_csv_from_excel(self,excel_file):
       
        csv_file = 'output_from_excel.csv'
        csv_output = StringIO()
        csvwriter = csv.writer(csv_output)
        header_written = False
    
        # Check the file extension to determine the library to use
        if excel_file.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
        elif excel_file.endswith('.xls'):
            with open('temp.xls', 'wb') as f:
                f.write(excel_file.getbuffer())
            wb = xlrd.open_workbook('temp.xls')  # Use xlrd for .xls

        # Iterate through the sheets in the workbook
        if excel_file.endswith('.xlsx'):
            for sheet in wb.sheetnames:  # Get sheet names for .xlsx
                ws = wb[sheet]
                
                
                # Extract the data similar to PDF extraction logic
                lines = []
                for row in ws.iter_rows(values_only=True):  # Use values_only to get cell values
                    lines.append(' '.join(map(str, row))) 
            
                day_month = lines[1] if len(lines) > 1 else ''  # Assuming the second line is day_month
                
                table_started = False
                
                for line in lines :
                    line_list =[]
                    if line:
                    
                        if any(word in line for word in self.full_month_index):
                        
                            day_month = self.extract_month(line)
                            #print("MONTH",day_month)
                        if any(word in line for word in self.name_index):
                        
                        
                            meter , name = self.extract_name_and_meter_excel(line)
                            
                            
                            print(name)
                            if meter :
                                meter = meter.split('\n')[0]
                            #print(",METER",meter)
                            

                        if any(word in line for word in self.header_index):
                            if not header_written:
                                line_list.append(line)
                            # print(line_list)
                                split_data = line_list[0].split('None')
                                cleaned_data = [item.strip() for item in split_data if item.strip()]
                                headers_list = [item.replace('\n',' ') for item in cleaned_data]
                                header_row = ['Name', 'MeterID'] + [word for word in headers_list if word != 'None' ]
                                header_length = len(header_row)
                                csvwriter.writerow(header_row)
                                header_written = True
                            table_started = True

                        if table_started:
                            row = line.split()
                            
                            try:
                                p = int(row[0])  # Assuming the first element is a number (day)
                                if row[3:]:  # Ensure there are more columns after the first 3
                                    date = f"{self.month.get(day_month[0][0])}/{row[0]}/{day_month[0][1]}"
                                    row[0] = date
                                    row_data = [item for item in row if item != 'None']
                                    print("ROW DATA>",row_data)
                                    all_digits = all(item.isdigit() for item in row_data[1:])
                                    main_row = [name if name else '', meter if meter else ''] + [item for item in row if item != 'None']
                                    if not all_digits:
                                        csvwriter.writerow(main_row)
                            except ValueError:
                                continue  # Ignore rows that don't start with a number

        elif excel_file.name.endswith('.xls'):
            for sheet_index in range(wb.nsheets):  # Loop through sheets by index
                ws = wb.sheet_by_index(sheet_index)
                name, meter = None, None
                
                lines = []
                for row_idx in range(ws.nrows):
                    row = ws.row_values(row_idx)  # Get values for the entire row
                    lines.append(' '.join(map(str, row)))  # Convert each row to a string
                
                day_month = lines[1] if len(lines) > 1 else ''  # Assuming the second line is day_month
                
                table_started = False
                
                for line in lines:
                    if any(word in line for word in self.header_index):
                        if not header_written:
                            head  = line.split()
                            head.pop()
                            header_row = ['Name', 'MeterID'] + line.split()
                            header_length  =len(header_row)
                            csvwriter.writerow(header_row)
                            header_written = True
                        table_started = True

                    if table_started:
                        row = line.split()
                        print("row>>",row)
                        try:
                            #p = int(row[0])  # Assuming the first element is a number (day)
                            if row[3:]:  # Ensure there are more columns after the first 3
                                #date = f"{month.get(day_month.split()[0])}/{row[0]}/{day_month.split()[1]}"
                                row[0] = 'date'
                                row.pop()
                                csvwriter.writerow([name if name else '', meter if meter else ''] + row)
                        except :
                            continue  # Ignore rows that don't start with a number

        csv_output.seek(0)
        df =pd.read_csv(csv_output)
        print("DF>>>",df)
        return csv_output, csv_file
    
    def take_files(self,uploaded_file):
        if uploaded_file.name.endswith('.pdf'):
            output , output_file,df = self.parse_csv_from_pdf(uploaded_file)
        if uploaded_file.name.endswith('.xls'):
            response  =  self.convert_single_to_multi_sheet_xlsx(uploaded_file)
            if response['status']:
                output , output_file = self.parse_csv_from_excel(response['file'])
        if output and output_file:
            return output , output_file 

    def convert_targa_file(self,uploaded_file):
        csv_file = 'output.csv'
        #line_list = []
        csv_output = StringIO()
        buff = StringIO()
        df = pd.read_excel(uploaded_file)
        df['Production_Day'] = pd.to_datetime(df['Production_Day'], format='%Y%m%d').dt.strftime('%m-%d-%Y')

        csv_output = df.to_csv(buff,index=False)
        buff.seek(0)
        print("bbb",csv_output)
      
        return buff, csv_file 
        

                  
